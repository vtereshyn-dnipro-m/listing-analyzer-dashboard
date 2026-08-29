# -*- coding: utf-8 -*-
"""
tests/fixture_template.py — миниатюра шаблона Amazon flat file.

Настоящий шаблон весит 3.6 МБ и в репозиторий не кладётся, а тестам нужен
файл с той же раскладкой строк: 1–6 служебные, данные с 7-й. Живёт
отдельным модулем, потому что нужен двум наборам сразу — импортировать
один тест из другого нельзя, они исполняются как скрипты.

Колонки НАРОЧНО сдвинуты, а подпись частичного обновления на выдуманном
языке: так ловится хардкод номеров колонок и испанской подписи.
"""
from __future__ import annotations

import io

import openpyxl

# --- миниатюра шаблона: та же раскладка строк, колонки НАРОЧНО сдвинуты
MP = "A1RKKUPIHCS9HS"
ITEM_NAME = f"item_name[marketplace_id={MP}][language_tag=es_ES]"
# «частичное» на выдуманном языке — чтобы поймать хардкод испанской подписи
PARTIAL = "Zzedit parcialny"
ATTRS = {2: "::listing_status", 4: "contribution_sku#1.value",
         5: "product_type#1.value", 6: "::record_action",
         11: f"{ITEM_NAME}#1.value", 13: "amzn1.volt.ca.product_id_value"}
DATA = [
    ("Activo", "17586000", "ABRASIVE_WHEELS", PARTIAL, "Disco viejo", "B0AAA"),
    ("Activo", "17586000-FBA", "ABRASIVE_WHEELS", PARTIAL, "Disco viejo", "B0AAA"),
    ("Activo", "S1_72321000", "DRILL", PARTIAL, "Taladro viejo", "B0BBB"),
    ("Activo", "80588000-FBA", "SANDER", PARTIAL, "Lijadora vieja", "B0CCC"),
    ("Eliminado", "99000000", "DRILL", PARTIAL, "Muerto", "B0DDD"),
]


def make_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    ws.cell(1, 1, "settings=feedType=256&templateIdentifier=TEST-1&dataRow=7")
    ws.cell(2, 1, "Instrucción")
    ws.cell(3, 1, "Reference Group")
    ws.cell(4, 4, "SKU")
    for col, name in ATTRS.items():
        ws.cell(5, col, name)
    ws.cell(6, 4, "ABC123")
    for i, row in enumerate(DATA):
        r = 7 + i
        for col, v in zip((2, 4, 5, 6, 11, 13), row):
            ws.cell(r, col, v)

    m = wb.create_sheet("AttributePTDMAP")
    for j, pt in enumerate(("ABRASIVE_WHEELS", "DRILL", "SANDER"), 1):
        m.cell(1, j, pt)
    m.cell(2, 1, "какая-то строка карты")

    d = wb.create_sheet("Dropdown Lists")
    d.cell(3, 3, "::record_action")
    d.cell(4, 3, "Zzcrear completo")
    d.cell(5, 3, PARTIAL)
    d.cell(6, 3, "Zzborrar")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build() -> bytes:
    """Байты миниатюры — то же, что вернул бы настоящий шаблон."""
    return make_template()
