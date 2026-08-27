# -*- coding: utf-8 -*-
"""
tests/test_flatfile_export.py — выгрузка обязана быть настоящим файлом Amazon.

Прежний экспорт из трёх колонок (sku / product_name / PartialUpdate) Кабинет
не принимал вовсе, и узналось это только когда файл попробовали загрузить.
Проверки здесь стерегут то, что ломается тихо:

  · служебные строки 1–6 должны дойти до готового файла без единой правки —
    в первой лежит settings с templateIdentifier, без неё файл чужой;
  · колонки ищутся по машинному имени из строки 5, а не по номеру: Amazon
    двигает их между выпусками шаблона;
  · подпись частичного обновления берётся из самого шаблона — она на языке
    кабинета, хардкод «Editar (actualización parcial)» сломался бы на любом
    другом маркетплейсе;
  · строка с типом товара, которого шаблон не покрывает, обязана попасть
    в список проблем, а не потеряться: типов 60, а шаблон принимает 30.

Настоящий шаблон весит 3.6 МБ и в репозиторий не кладётся, поэтому здесь
собирается его миниатюра с той же раскладкой строк.

Запуск (pytest не нужен):  python tests/test_flatfile_export.py
"""
from __future__ import annotations

import io
import pathlib
import sys

import openpyxl
import pandas as pd

ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import services.db                                     # noqa: E402
services.db.get_conn = lambda: type(
    "C", (), {"close": lambda self: None})()
pd.read_sql = lambda *a, **k: pd.DataFrame()

import services.flatfile as ff                         # noqa: E402
from services.flatfile_template import (               # noqa: E402
    parse_template, build_file, sku_for, type_index)

FAILS: list[str] = []


def check(name: str, cond: bool) -> None:
    print(("  OK   " if cond else "  FAIL ") + name)
    if not cond:
        FAILS.append(name)


# --- миниатюра шаблона: та же раскладка строк, колонки НАРОЧНО сдвинуты
MP = "A1RKKUPIHCS9HS"
ITEM_NAME = f"item_name[marketplace_id={MP}][language_tag=es_ES]"
# «частичное» на выдуманном языке — чтобы поймать хардкод испанской подписи
PARTIAL = "Zzedit parcialny"
ATTRS = {2: "::listing_status", 4: "contribution_sku#1.value",
         5: "product_type#1.value", 6: "::record_action",
         11: f"{ITEM_NAME}#1.value", 13: "amzn1.volt.ca.product_id_value"}
DATA = [
    ("Activo", "17586000", "ABRASIVE_WHEELS", PARTIAL, "Disco viejo", "B0AAA"),
    ("Activo", "17586000-FBA", "ABRASIVE_WHEELS", PARTIAL, "Disco viejo", "B0AAA"),
    ("Activo", "S1_72321000", "DRILL", PARTIAL, "Taladro viejo", "B0BBB"),
    ("Activo", "80588000-FBA", "SANDER", PARTIAL, "Lijadora vieja", "B0CCC"),
    ("Eliminado", "99000000", "DRILL", PARTIAL, "Muerto", "B0DDD"),
]


def make_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    ws.cell(1, 1, "settings=feedType=256&templateIdentifier=TEST-1&dataRow=7")
    ws.cell(2, 1, "Instrucción")
    ws.cell(3, 1, "Reference Group")
    ws.cell(4, 4, "SKU")
    for col, name in ATTRS.items():
        ws.cell(5, col, name)
    ws.cell(6, 4, "ABC123")
    for i, row in enumerate(DATA):
        r = 7 + i
        for col, v in zip((2, 4, 5, 6, 11, 13), row):
            ws.cell(r, col, v)

    m = wb.create_sheet("AttributePTDMAP")
    for j, pt in enumerate(("ABRASIVE_WHEELS", "DRILL", "SANDER"), 1):
        m.cell(1, j, pt)
    m.cell(2, 1, "какая-то строка карты")

    d = wb.create_sheet("Dropdown Lists")
    d.cell(3, 3, "::record_action")
    d.cell(4, 3, "Zzcrear completo")
    d.cell(5, 3, PARTIAL)
    d.cell(6, 3, "Zzborrar")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


RAW = make_template()
tpl = parse_template("0_TEST-RANGE.xlsx", RAW)

# --- разбор
check("slot взят из имени файла", tpl["slot"] == "0_TEST-RANGE")
check("колонки найдены по имени атрибута, а не по номеру",
      tpl["columns"]["contribution_sku#1.value"] == 4
      and tpl["columns"]["product_type#1.value"] == 5
      and tpl["columns"]["::record_action"] == 6
      and tpl["columns"]["item_name"] == 11)
check("item_name опознан вместе с marketplace_id и language_tag",
      tpl["item_name_attr"] == f"{ITEM_NAME}#1.value")
check("подпись частичного обновления взята из шаблона, а не захардкожена",
      tpl["partial_label"] == PARTIAL)
check("типы товара взяты с листа AttributePTDMAP",
      tpl["product_types"] == ["ABRASIVE_WHEELS", "DRILL", "SANDER"])
check("карта ASIN собрана", len(tpl["sku_map"]) == 4 and tpl["rows_seen"] == 5)

# --- эталон: строк данных в нём быть не должно
ws_ref = openpyxl.load_workbook(
    io.BytesIO(tpl["template_bytes"]), read_only=True)["Plantilla"]
check("в эталоне вырезаны чужие строки данных", ws_ref.max_row <= 6)
check("служебные строки в эталоне остались",
      ws_ref.cell(5, 4).value == "contribution_sku#1.value"
      if ws_ref.max_row >= 5 else False)

# --- выбор SKU
check("при паре FBM/FBA берётся FBM",
      sku_for([tpl], "B0AAA") == ("17586000", "ABRASIVE_WHEELS", "template"))
check("префикс набора S1_ сохраняется",
      sku_for([tpl], "B0BBB")[0] == "S1_72321000")
check("если строка только -FBA, берём её и помечаем источник",
      sku_for([tpl], "B0CCC") == ("80588000-FBA", "SANDER", "template-fba"))
check("удалённый листинг не выдаёт SKU молча",
      sku_for([tpl], "B0DDD")[0] == "99000000")
check("незнакомый ASIN не выдумывает SKU", sku_for([tpl], "B0ZZZ") == ("", "", ""))
check("индекс типов покрывает все три", len(type_index([tpl])) == 3)

# --- сборка
ROWS = [{"sku": "17586000", "product_type": "ABRASIVE_WHEELS",
         "title": 'Disco «ULTRA» 125 mm & <10> — nuevo'},
        {"sku": "S1_72321000", "product_type": "DRILL",
         "title": "Taladro nuevo"}]
built = build_file(tpl, ROWS)
wb_out = openpyxl.load_workbook(io.BytesIO(built), read_only=True,
                                data_only=True)
wb_in = openpyxl.load_workbook(io.BytesIO(RAW), read_only=True, data_only=True)
out, src = wb_out["Plantilla"], wb_in["Plantilla"]
head_out = [r for r in out.iter_rows(min_row=1, max_row=6, values_only=True)]
head_src = [r for r in src.iter_rows(min_row=1, max_row=6, values_only=True)]
check("строки 1–6 дошли без изменений", head_out == head_src)
check("settings из первой строки на месте",
      str(head_out[0][0]).startswith("settings=feedType=256"))
check("все листы шаблона сохранены",
      wb_out.sheetnames == wb_in.sheetnames)

body = [r for r in out.iter_rows(min_row=7, values_only=True)]
check("наших строк ровно столько, сколько дали", len(body) == 2)
r7 = body[0]
check("SKU, тип и действие легли в свои колонки",
      r7[3] == "17586000" and r7[4] == "ABRASIVE_WHEELS" and r7[5] == PARTIAL)
check("тайтл лёг в item_name", r7[10] == 'Disco «ULTRA» 125 mm & <10> — nuevo')
check("остальные колонки пустые — частичное обновление их не тронет",
      r7[0] is None and r7[1] is None and r7[12] is None)

# --- раскладка по шаблонам и список непопавших
ff.templates_for = lambda mp: [tpl] if mp == "es" else []
ff.load_product_types = lambda: {("B0EEE", "es"): "WRENCH"}
ff.load_sku_map = lambda: {("B0EEE", "es"): ("55000000", "catalog")}
ACC = pd.DataFrame([
    dict(asin="B0AAA", marketplace="es", after_title="Nuevo A"),
    dict(asin="B0BBB", marketplace="es", after_title="Nuevo B"),
    dict(asin="B0EEE", marketplace="es", after_title="Nuevo E"),   # тип вне шаблона
    dict(asin="B0FFF", marketplace="es", after_title="Nuevo F"),   # ни SKU, ни типа
    dict(asin="B0GGG", marketplace="de", after_title="Neu G"),     # нет шаблона
])
plan, bad = ff.plan_export(ACC)
check("план — один файл на один шаблон", len(plan) == 1)
check("в файл попали только размещаемые строки",
      len(plan[0]["rows"]) == 2)
reasons = {b["asin"]: b["reason"] for b in bad}
check("тип вне шаблонов назван причиной",
      reasons.get("B0EEE") == "type_unknown")
check("отсутствие SKU названо причиной", reasons.get("B0FFF") == "no_sku")
check("маркетплейс без шаблона назван причиной",
      reasons.get("B0GGG") == "no_template")
check("ничего не потеряно: сумма сходится",
      len(plan[0]["rows"]) + len(bad) == len(ACC))

name, mime, data = ff.build_flat_export(plan, "2026-08-27")
check("имя файла ведёт к шаблону-источнику и хранит его расширение",
      name == "0_TEST-RANGE_es_2026-08-27_2.xlsx")
check("mime соответствует расширению", mime == ff.XLSX_MIME)
check("отдаётся непустой файл", len(data) > 3000)

print()
print("ИТОГ:", "все проверки прошли" if not FAILS
      else f"{len(FAILS)} провалов: {FAILS}")
sys.exit(1 if FAILS else 0)
