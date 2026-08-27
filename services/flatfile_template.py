# -*- coding: utf-8 -*-
"""
services/flatfile_template.py — шаблон Amazon flat file как эталон.

Amazon принимает не произвольную таблицу, а СВОЙ файл: вкладка «Plantilla»,
строки 1–6 служебные (в строке 1 — settings с templateIdentifier, в строке 5 —
машинные имена атрибутов), данные с 7-й. Наш прежний экспорт из трёх колонок
(sku / product_name / PartialUpdate) такой загрузкой не принимается.

Отсюда устройство модуля: файл-эталон приходит от человека один раз
(Seller Central → Category Listings Report), у него вырезаются строки данных,
и дальше при каждой выгрузке мы дописываем свои строки в НЕТРОНУТУЮ копию.
Запись сделана хирургией по zip: подменяется ровно один XML листа, все
остальные части архива копируются байт в байт. Так переживают и settings,
и выпадающие списки, и картинки — то, что openpyxl при пересохранении теряет.

Три вещи, ради которых файл читается целиком, а не только строки 1–6:

1. **Типов товара 60, а шаблон покрывает 30.** Два файла отчёта делят их
   без пересечений (ABRASIVE_WHEELS…LEVEL и MACHINE_LUBRICANT…WRENCH), и
   строка с DRILL, положенная во второй файл, не пройдёт валидацию. Список
   покрываемых типов берётся с листа AttributePTDMAP.
2. **Карта ASIN → SKU → product_type.** Отчёт — выгрузка самого Кабинета,
   в нём настоящие продавцовые SKU (включая наборы S1_/S2_ и парные -FBA)
   и настоящий product_type. Это надёжнее наших источников.
3. **Подпись частичного обновления зависит от языка шаблона.** Берётся
   с листа Dropdown Lists: у колонки `::record_action` три значения в
   фиксированном порядке — полное обновление, частичное, удаление.
"""
from __future__ import annotations

import io
import json
import re
import zipfile

import pandas as pd
import streamlit as st

from services.db import get_conn

SHEET = "Plantilla"
MAP_SHEET = "AttributePTDMAP"
LIST_SHEET = "Dropdown Lists"
ATTR_ROW = 5          # машинные имена атрибутов
DATA_ROW = 7          # с этой строки идут данные
KEEP_ROWS = 6         # строки 1..6 — служебные, переносятся как есть

# машинные имена колонок, которые заполняем
A_SKU = "contribution_sku#1.value"
A_TYPE = "product_type#1.value"
A_ACTION = "::record_action"
A_STATUS = "::listing_status"
A_ITEM_NAME = "item_name["            # дальше [marketplace_id=…][language_tag=…]
A_ID_VALUE = "amzn1.volt.ca.product_id_value"

# запасная подпись, если лист со списками не прочитался (es_ES)
PARTIAL_FALLBACK = "Editar (actualización parcial)"

ROW_RE = re.compile(r"<row\b[^>]*?(?:/>|>.*?</row>)", re.S)
ROW_N_RE = re.compile(r'\br="(\d+)"')
CELL_RE = re.compile(r"<c\b[^>]*?(?:/>|>.*?</c>)", re.S)
CELL_REF_RE = re.compile(r'\br="([A-Z]+)(\d+)"')
CELL_S_RE = re.compile(r'\bs="(\d+)"')
DIM_RE = re.compile(r'<dimension ref="[^"]*"\s*/>')


def col_letter(n: int) -> str:
    """1 → A, 27 → AA."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _esc(v) -> str:
    return (str(v).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


def _row_num(row_xml: str) -> int:
    m = ROW_N_RE.search(row_xml)
    return int(m.group(1)) if m else 0


# ---------------------------------------------------------------- разбор

def _sheet_path(z: zipfile.ZipFile, name: str) -> str:
    """Путь к XML листа по его человеческому имени."""
    wb = z.read("xl/workbook.xml").decode("utf-8")
    m = re.search(r'<sheet[^>]*name="%s"[^>]*?r:id="(rId\d+)"' % re.escape(name),
                  wb)
    if not m:
        m = re.search(r'<sheet[^>]*r:id="(rId\d+)"[^>]*?name="%s"'
                      % re.escape(name), wb)
    if not m:
        raise ValueError(f"нет листа {name}")
    # порядок атрибутов в rels у Excel и openpyxl разный — ищем оба
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid = m.group(1)
    mt = (re.search(r'<Relationship[^>]*Id="%s"[^>]*?Target="([^"]+)"' % rid, rels)
          or re.search(r'<Relationship[^>]*Target="([^"]+)"[^>]*?Id="%s"' % rid, rels))
    if not mt:
        raise ValueError(f"нет связи для листа {name}")
    tgt = mt.group(1).lstrip("/")
    return tgt if tgt.startswith("xl/") else "xl/" + tgt


def _partial_label(wb) -> str:
    """Подпись «частичное обновление» на языке шаблона.

    На листе Dropdown Lists у колонки `::record_action` три значения подряд
    в порядке Amazon: полное обновление, частичное, удаление. Берём второе —
    так подпись не приходится хардкодить под каждый язык кабинета.
    """
    try:
        ws = wb[LIST_SHEET]
        col = None
        for r in ws.iter_rows(min_row=1, max_row=4, values_only=True):
            for j, v in enumerate(r):
                if v and str(v).strip() == A_ACTION:
                    col = j
                    break
            if col is not None:
                break
        if col is None:
            return PARTIAL_FALLBACK
        vals = []
        for r in ws.iter_rows(min_row=1, values_only=True):
            v = r[col] if col < len(r) else None
            if v and str(v).strip() and str(v).strip() != A_ACTION:
                vals.append(str(v).strip())
            if len(vals) >= 3:
                break
        return vals[1] if len(vals) >= 2 else PARTIAL_FALLBACK
    except Exception:
        return PARTIAL_FALLBACK


def _covered_types(wb) -> list[str]:
    """Типы товара, которые покрывает шаблон — шапка листа AttributePTDMAP."""
    try:
        ws = wb[MAP_SHEET]
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        return sorted({str(v).strip() for v in hdr
                       if v and re.fullmatch(r"[A-Z][A-Z_0-9]{2,}", str(v).strip())})
    except Exception:
        return []


def parse_template(file_name: str, data: bytes) -> dict:
    """Разбор загруженного файла. Бросает ValueError с внятным текстом."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                    data_only=True)
    except Exception as e:
        raise ValueError(f"файл не читается как Excel: {e}") from e
    if SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"нет вкладки «{SHEET}» — это не шаблон Amazon")

    ws = wb[SHEET]
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=DATA_ROW - 1,
                                       values_only=True), 1):
        rows.append(r)
    if len(rows) < ATTR_ROW:
        wb.close()
        raise ValueError("во вкладке «Plantilla» меньше строк, чем у шаблона")

    # строка 5 → номер колонки для каждого машинного имени
    attrs = rows[ATTR_ROW - 1]
    cols: dict[str, int] = {}
    for j, v in enumerate(attrs, 1):
        if v and str(v).strip():
            cols.setdefault(str(v).strip(), j)
    item_name = next((a for a in cols if a.startswith(A_ITEM_NAME)), None)
    missing = [n for n, ok in (("SKU", A_SKU in cols),
                               ("product_type", A_TYPE in cols),
                               ("::record_action", A_ACTION in cols),
                               ("item_name", item_name is not None)) if not ok]
    if missing:
        wb.close()
        raise ValueError("в строке 5 не найдены колонки: " + ", ".join(missing))

    # карта из строк данных: ASIN → строки Кабинета
    c_sku, c_type = cols[A_SKU], cols[A_TYPE]
    c_id = cols.get(A_ID_VALUE)
    c_st = cols.get(A_STATUS)
    width = max(c_sku, c_type, c_id or 0, c_st or 0)
    sku_map: dict[str, list] = {}
    if c_id:
        for r in ws.iter_rows(min_row=DATA_ROW, max_col=width, values_only=True):
            asin = r[c_id - 1]
            sku = r[c_sku - 1]
            if not asin or not sku:
                continue
            sku_map.setdefault(str(asin).strip(), []).append({
                "sku": str(sku).strip(),
                "product_type": str(r[c_type - 1] or "").strip(),
                "status": str(r[c_st - 1] or "").strip() if c_st else "",
            })

    out = {
        "file_name": file_name,
        "slot": re.sub(r"\.(xlsm|xlsx)$", "", file_name, flags=re.I),
        "sheet_path": "",
        "columns": {A_SKU: c_sku, A_TYPE: c_type, A_ACTION: cols[A_ACTION],
                    "item_name": cols[item_name]},
        "item_name_attr": item_name,
        "partial_label": _partial_label(wb),
        "product_types": _covered_types(wb),
        "sku_map": sku_map,
        "rows_seen": sum(len(v) for v in sku_map.values()),
    }
    wb.close()

    z = zipfile.ZipFile(io.BytesIO(data))
    out["sheet_path"] = _sheet_path(z, SHEET)
    if not out["product_types"]:
        # без карты типов не понять, какие товары шаблон принимает;
        # берём типы из его же строк данных — беднее, но работает
        out["product_types"] = sorted(
            {x["product_type"] for v in sku_map.values() for x in v
             if x["product_type"]})
    out["template_bytes"], out["styles"] = strip_data_rows(
        data, out["sheet_path"])
    return out


# ---------------------------------------------------------------- запись

def strip_data_rows(data: bytes, sheet_path: str) -> tuple[bytes, dict]:
    """Эталон: строки 1–6 на месте, строки данных вырезаны.

    Заодно снимает со старой строки 7 оформление ячеек — своим строкам мы
    отдадим те же стили, иначе выгрузка выглядит чужой в файле Amazon.
    """
    z = zipfile.ZipFile(io.BytesIO(data))
    xml = z.read(sheet_path).decode("utf-8")
    a, b = xml.index("<sheetData"), xml.index("</sheetData>")
    head, body, tail = xml[:a], xml[a:b], xml[b:]
    open_tag = body[:body.index(">") + 1]

    keep, styles = [], {}
    for row in ROW_RE.findall(body):
        n = _row_num(row)
        if n <= KEEP_ROWS:
            keep.append(row)
        elif n == DATA_ROW and not styles:
            for c in CELL_RE.findall(row):
                cm, sm = CELL_REF_RE.search(c), CELL_S_RE.search(c)
                if cm and sm:
                    styles[cm.group(1)] = sm.group(1)
    new_xml = (_fix_dim(head, KEEP_ROWS) + open_tag + "".join(keep) + tail)
    return _repack(data, {sheet_path: new_xml.encode("utf-8")}), styles


def _fix_dim(head: str, last_row: int) -> str:
    return DIM_RE.sub(lambda m: re.sub(r"(\d+)(?=\"\s*/>$)", str(last_row),
                                       m.group(0)), head)


def _repack(src: bytes, replace: dict) -> bytes:
    """Копия архива, где подменены только указанные части."""
    zin = zipfile.ZipFile(io.BytesIO(src))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zo:
        for it in zin.infolist():
            zi = zipfile.ZipInfo(it.filename, date_time=it.date_time)
            zi.compress_type = it.compress_type
            zi.external_attr = it.external_attr
            zo.writestr(zi, replace.get(it.filename) or zin.read(it.filename))
    return buf.getvalue()


def build_file(tpl: dict, rows: list[dict]) -> bytes:
    """Готовый файл: эталон + наши строки с 7-й.

    Заполняются только SKU, product_type, ::record_action и item_name.
    Остальные колонки пустые намеренно: при частичном обновлении Amazon
    не трогает то, чего нет в файле, — так правка тайтла не может задеть
    цену, картинки или атрибуты.
    """
    data = tpl["template_bytes"]
    path, cols, styles = tpl["sheet_path"], tpl["columns"], tpl.get("styles") or {}
    z = zipfile.ZipFile(io.BytesIO(data))
    xml = z.read(path).decode("utf-8")
    a, b = xml.index("<sheetData"), xml.index("</sheetData>")
    head, body, tail = xml[:a], xml[a:b], xml[b:]
    open_tag = body[:body.index(">") + 1]
    keep = [r for r in ROW_RE.findall(body) if _row_num(r) <= KEEP_ROWS]

    out = []
    for i, r in enumerate(rows):
        n = DATA_ROW + i
        vals = {cols[A_SKU]: r["sku"], cols[A_TYPE]: r["product_type"],
                cols[A_ACTION]: tpl["partial_label"],
                cols["item_name"]: r["title"]}
        cells = []
        for ci in sorted(vals):
            letter = col_letter(ci)
            style = f' s="{styles[letter]}"' if letter in styles else ""
            cells.append(f'<c r="{letter}{n}"{style} t="inlineStr">'
                         f'<is><t xml:space="preserve">{_esc(vals[ci])}</t>'
                         f'</is></c>')
        out.append(f'<row r="{n}">' + "".join(cells) + "</row>")

    new_xml = (_fix_dim(head, KEEP_ROWS + len(rows)) + open_tag
               + "".join(keep) + "".join(out) + tail)
    return _repack(data, {path: new_xml.encode("utf-8")})


# ---------------------------------------------------------------- хранение

SESSION_KEY = "flatfile.templates"      # запасной склад, пока нет таблицы
_JSON_FIELDS = ("columns", "product_types", "sku_map", "styles")


def _pack(tpl: dict) -> dict:
    out = dict(tpl)
    for f in _JSON_FIELDS:
        out[f] = json.dumps(tpl.get(f), ensure_ascii=False)
    return out


def _unpack(row: dict) -> dict:
    out = dict(row)
    for f in _JSON_FIELDS:
        v = out.get(f)
        out[f] = json.loads(v) if isinstance(v, str) else (v or {})
    # psycopg2 отдаёт bytea как memoryview — дальше он идёт в zipfile
    b = out.get("template_bytes")
    out["template_bytes"] = bytes(b) if b is not None else b""
    out["columns"] = {k: int(v) for k, v in (out.get("columns") or {}).items()}
    return out


def save_template(marketplace: str, tpl: dict) -> str | None:
    """В Lakebase; при неудаче — в сессию. Возвращает текст ошибки или None."""
    p = _pack(tpl)
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO flatfile_templates
                    (marketplace, slot, file_name, sheet_path, columns,
                     item_name_attr, partial_label, product_types, sku_map,
                     styles, rows_seen, template_bytes, uploaded_at)
                VALUES (%(mp)s, %(slot)s, %(fn)s, %(sp)s, %(cols)s::jsonb,
                        %(ina)s, %(pl)s, %(pt)s::jsonb, %(sm)s::jsonb,
                        %(stl)s::jsonb, %(rs)s, %(tb)s, now())
                ON CONFLICT (marketplace, slot) DO UPDATE SET
                    file_name = EXCLUDED.file_name,
                    sheet_path = EXCLUDED.sheet_path,
                    columns = EXCLUDED.columns,
                    item_name_attr = EXCLUDED.item_name_attr,
                    partial_label = EXCLUDED.partial_label,
                    product_types = EXCLUDED.product_types,
                    sku_map = EXCLUDED.sku_map,
                    styles = EXCLUDED.styles,
                    rows_seen = EXCLUDED.rows_seen,
                    template_bytes = EXCLUDED.template_bytes,
                    uploaded_at = now()
                """,
                {"mp": marketplace, "slot": p["slot"], "fn": p["file_name"],
                 "sp": p["sheet_path"], "cols": p["columns"],
                 "ina": p["item_name_attr"], "pl": p["partial_label"],
                 "pt": p["product_types"], "sm": p["sku_map"],
                 "stl": p["styles"], "rs": int(p["rows_seen"]),
                 "tb": p["template_bytes"]})
        conn.commit()
        conn.close()
        load_templates.clear()
        return None
    except Exception as e:
        # склад в сессии — чтобы выгрузка работала уже сейчас, до миграции;
        # ошибка при этом не прячется, страница её показывает
        st.session_state.setdefault(SESSION_KEY, {})[
            (marketplace, tpl["slot"])] = dict(tpl, stored="session")
        return str(e).strip().splitlines()[0] if str(e).strip() else repr(e)


@st.cache_data(ttl=120, show_spinner=False)
def load_templates(marketplace: str) -> list[dict]:
    """Шаблоны маркетплейса из Lakebase (без сессионных — они добавляются
    в templates_for(), потому что кэш их бы заморозил)."""
    try:
        conn = get_conn()
        df = pd.read_sql(
            """
            SELECT marketplace, slot, file_name, sheet_path, columns,
                   item_name_attr, partial_label, product_types, sku_map,
                   styles, rows_seen, template_bytes, uploaded_at
            FROM flatfile_templates
            WHERE marketplace = %(mp)s
            ORDER BY slot
            """, conn, params={"mp": str(marketplace).lower()})
        conn.close()
    except Exception:
        return []
    return [_unpack(r) for r in df.to_dict("records")]


def templates_for(marketplace: str) -> list[dict]:
    """Шаблоны маркетплейса: из базы плюс загруженные в этой сессии."""
    mp = str(marketplace).lower()
    out = {t["slot"]: dict(t, stored="db") for t in load_templates(mp)}
    for (m, slot), tpl in (st.session_state.get(SESSION_KEY) or {}).items():
        if m == mp:
            out.setdefault(slot, tpl)
    return [out[k] for k in sorted(out)]


def type_index(templates: list[dict]) -> dict:
    """product_type → шаблон, который его принимает."""
    idx = {}
    for tpl in templates:
        for pt in tpl.get("product_types") or []:
            idx.setdefault(pt, tpl)
    return idx


def sku_for(templates: list[dict], asin: str) -> tuple[str, str, str]:
    """(sku, product_type, источник) по ASIN из отчёта Кабинета.

    Правило SKU: обновляем FBM-листинг, поэтому строка с суффиксом -FBA
    берётся только когда другой нет (у пары товаров она единственная).
    Префиксы наборов S1_/S2_ и прочие суффиксы не трогаем — это настоящие
    SKU продавца, а не наша разметка.
    """
    for tpl in templates:
        rows = (tpl.get("sku_map") or {}).get(str(asin))
        if not rows:
            continue
        live = [r for r in rows if r.get("status", "").lower() != "eliminado"]
        pool = live or rows
        fbm = [r for r in pool if not r["sku"].upper().endswith("-FBA")]
        pick = (fbm or pool)[0]
        src = "template" if fbm else "template-fba"
        return pick["sku"], pick.get("product_type", ""), src
    return "", "", ""
